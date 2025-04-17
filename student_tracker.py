from openpyxl import load_workbook
workbook = load_workbook('D:/student_records.xlsx')
sheet = workbook.active
next_row=sheet.max_row+1
id=int(input("Enter student id: "))
name=input("Enter student name: ")
marks=int(input("Enter student marks : "))
sheet.cell(row=next_row,column=1).value=id
sheet.cell(row=next_row,column=2).value=name
sheet.cell(row=next_row,column=3).value=marks
workbook.save('D:/student_records.xlsx')
print("added successfully")
print("current student record:\n") 
for row in sheet.iter_rows(values_only=True):
    print(row)
